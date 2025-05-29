import flet as ft
from GestorPrincipal import GestorCedulas
from IR_scanner import EscanerCedula
import base64
import pygame
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

pygame.init()
pygame.mixer.init()

gestor = GestorCedulas()

class CedulaApp:
    def __init__(self):
        self.escaner = None
        self.cedula_encontrada = None
        self.showing_img = None
        self.status_text = None
        self.cedula_text = None
        self.main_page = None
        self.cedula_dialog = None
        self.page = None
        self.detenido = False

        # Campos para el formulario
        self.nombre_input = None
        self.cedula_input = None
        self.especialidad_input = None
        self.anio_input = None
        self.seccion_input = None
        self.buscar_cedula_input = None
        self.resultado_container = None
        self.tabla_registros = None
        self.mensaje_estado = None
        self.escaner = None

    def iniciar_app(self, page: ft.Page):
        self.page = page
        page.title = "Escáner de Cédula"
        page.theme_mode = ft.ThemeMode.LIGHT
        page.padding = 24
        page.window_width = 1024
        page.window_height = 768
        page.bgcolor = ft.Colors.GREY_50
        
        # Crear componentes de UI modernizados
        self.img_view = ft.Container(
            content=ft.Image(
                src="assets/desactivar-camara.png",
                width=640,
                height=480,
                fit=ft.ImageFit.CONTAIN,
            ),
            border_radius=16,
            bgcolor=ft.Colors.WHITE,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=12,
                color=ft.Colors.with_opacity(0.1, ft.Colors.BLACK),
                offset=ft.Offset(0, 4),
            ),
            padding=16,
        )

        
        self.status_text = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.RADIO_BUTTON_UNCHECKED, size=16, color=ft.Colors.BLUE_600),
                ft.Text(
                    value="Estado: Esperando...",
                    size=14,
                    color=ft.Colors.GREY_700,
                    weight=ft.FontWeight.W_500,
                )
            ], tight=True),
            padding=ft.padding.symmetric(horizontal=12, vertical=8),
            border_radius=8,
            bgcolor=ft.Colors.BLUE_50,
            border=ft.border.all(1, ft.Colors.BLUE_100),
        )
        
        self.cedula_text = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.BADGE_OUTLINED, size=24, color=ft.Colors.GREEN_600),
                ft.Text(
                    value="Cédula: No detectada",
                    size=18,
                    weight=ft.FontWeight.W_600,
                    color=ft.Colors.GREY_800,
                )
            ], tight=True),
            padding=ft.padding.all(16),
            border_radius=12,
            bgcolor=ft.Colors.WHITE,
            border=ft.border.all(1, ft.Colors.GREY_200),
            shadow=ft.BoxShadow(
                spread_radius=0,
                blur_radius=4,
                color=ft.Colors.with_opacity(0.05, ft.Colors.BLACK),
                offset=ft.Offset(0, 2),
            ),
        )
        
        # Botones de control modernizados
        btn_iniciar = ft.ElevatedButton(
            text="Iniciar Escaneo",
            icon=ft.Icons.CAMERA_ALT_OUTLINED,
            on_click=self.on_iniciar_escaneo,
            style=ft.ButtonStyle(
                color=ft.Colors.WHITE,
                bgcolor=ft.Colors.GREEN_600,
                padding=ft.padding.symmetric(horizontal=24, vertical=16),
                shape=ft.RoundedRectangleBorder(radius=12),
                elevation=3,
                text_style=ft.TextStyle(size=16, weight=ft.FontWeight.W_500),
            ),
            height=56,
            width=180,
        )
        
        btn_detener = ft.ElevatedButton(
            text="Detener Escaneo",
            icon=ft.Icons.STOP_CIRCLE_OUTLINED,
            on_click=self.on_detener_escaneo,
            style=ft.ButtonStyle(
                color=ft.Colors.WHITE,
                bgcolor=ft.Colors.RED_500,
                padding=ft.padding.symmetric(horizontal=24, vertical=16),
                shape=ft.RoundedRectangleBorder(radius=12),
                elevation=3,
                text_style=ft.TextStyle(size=16, weight=ft.FontWeight.W_500),
            ),
            height=56,
            width=180,
        )
        
        self.inicializar_componentes_gestion()
        self.cargar_historial()
        
        # Pestaña de escaneo modernizada
        tab_escaneo = ft.Tab(
            text="Escaneo",
            icon=ft.Icons.DOCUMENT_SCANNER_OUTLINED,
            content=ft.Container(
                content=ft.Column(
                    [
                        ft.Container(
                            content=ft.Text(
                                "Escáner de Cédulas",
                                size=24,
                                weight=ft.FontWeight.W_600,
                                color=ft.Colors.GREY_800,
                            ),
                            margin=ft.margin.only(bottom=24),
                        ),
                        ft.Row(
                            [btn_iniciar, btn_detener],
                            alignment=ft.MainAxisAlignment.CENTER,
                            spacing=16,
                        ),
                        ft.Container(height=24),
                        self.status_text,
                        ft.Container(height=16),
                        self.cedula_text,
                        ft.Container(height=24),
                        self.img_view,
                    ],
                    alignment=ft.MainAxisAlignment.START,
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    spacing=0,
                ),
                padding=24,
                bgcolor=ft.Colors.WHITE,
                border_radius=16,
                margin=ft.margin.all(8),
                shadow=ft.BoxShadow(
                    spread_radius=0,
                    blur_radius=8,
                    color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                    offset=ft.Offset(0, 2),
                ),
            ),
        )
        
        # Pestaña de crear registro modernizada
        tab_crear = ft.Tab(
            text="Crear Registro",
            icon=ft.Icons.ADD_CIRCLE_OUTLINE,
            content=ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.PERSON_ADD_OUTLINED, size=28, color=ft.Colors.BLUE_600),
                            ft.Text(
                                "Nuevo Registro de Estudiante", 
                                size=24, 
                                weight=ft.FontWeight.W_600,
                                color=ft.Colors.GREY_800,
                            )
                        ], tight=True),
                        margin=ft.margin.only(bottom=32),
                    ),
                    ft.Container(
                        content=ft.Column([
                            self.nombre_input,
                            ft.Container(height=16),
                            self.cedula_input,
                            ft.Container(height=16),
                            self.especialidad_input,
                            ft.Container(height=16),
                            self.anio_input,
                            ft.Container(height=16),
                            self.seccion_input,
                            ft.Container(height=32),
                            ft.ElevatedButton(
                                "Crear Registro",
                                icon=ft.Icons.SAVE_OUTLINED,
                                on_click=self.crear_click,
                                style=ft.ButtonStyle(
                                    color=ft.Colors.WHITE,
                                    bgcolor=ft.Colors.BLUE_600,
                                    padding=ft.padding.symmetric(horizontal=32, vertical=16),
                                    shape=ft.RoundedRectangleBorder(radius=12),
                                    elevation=3,
                                    text_style=ft.TextStyle(size=16, weight=ft.FontWeight.W_500),
                                ),
                                height=56,
                                width=200,
                            )
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                        padding=32,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=16,
                        shadow=ft.BoxShadow(
                            spread_radius=0,
                            blur_radius=8,
                            color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                            offset=ft.Offset(0, 2),
                        ),
                    )
                ]),
                padding=24,
                bgcolor=ft.Colors.GREY_50,
            )
        )

        # Pestaña de ver registros modernizada
        tab_listar = ft.Tab(
            text="Ver Registros",
            icon=ft.Icons.LIST_ALT_OUTLINED,
            content=ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.VIEW_LIST_OUTLINED, size=28, color=ft.Colors.GREEN_600),
                            ft.Text(
                                "Listado de Registros", 
                                size=24, 
                                weight=ft.FontWeight.W_600,
                                color=ft.Colors.GREY_800,
                            )
                        ], tight=True),
                        margin=ft.margin.only(bottom=24)
                        
                    ),
                    ft.Container(
                        content=ft.ElevatedButton(
                            "Refrescar Listado",
                            icon=ft.Icons.REFRESH,
                            on_click=lambda _: self.cargar_registros(),
                            style=ft.ButtonStyle(
                                color=ft.Colors.WHITE,
                                bgcolor=ft.Colors.GREEN_600,
                                padding=ft.padding.symmetric(horizontal=24, vertical=12),
                                shape=ft.RoundedRectangleBorder(radius=8),
                                elevation=2,
                            ),
                            height=48,
                        ),
                        margin=ft.margin.only(bottom=24),
                        alignment= ft.alignment.center
                    ),
                    self.registros_container
                ]),
                padding=24,
                bgcolor=ft.Colors.GREY_50,
            )
        )
        
        # Pestaña de buscar modernizada
        tab_buscar = ft.Tab(
            text="Buscar y Administrar",
            icon=ft.Icons.SEARCH_OUTLINED,
            content=ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.MANAGE_SEARCH_OUTLINED, size=28, color=ft.Colors.ORANGE_600),
                            ft.Text(
                                "Buscar y Administrar Registros", 
                                size=24, 
                                weight=ft.FontWeight.W_600,
                                color=ft.Colors.GREY_800,
                            )
                        ], tight=True),
                        margin=ft.margin.only(bottom=32),
                    ),
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                self.buscar_cedula_input,
                                ft.Container(width=16),
                                ft.ElevatedButton(
                                    "Buscar",
                                    icon=ft.Icons.SEARCH,
                                    on_click=self.buscar_click,
                                    style=ft.ButtonStyle(
                                        color=ft.Colors.WHITE,
                                        bgcolor=ft.Colors.BLUE_600,
                                        padding=ft.padding.symmetric(horizontal=24, vertical=12),
                                        shape=ft.RoundedRectangleBorder(radius=12),
                                        elevation=2,
                                    ),
                                    height=56,
                                    width=150,
                                )
                            ], alignment=ft.MainAxisAlignment.CENTER),
                            ft.Container(height=24),
                            self.resultado_container,
                            ft.Container(height=24),
                            ft.Row([
                                ft.ElevatedButton(
                                    "Eliminar Registro",
                                    icon=ft.Icons.DELETE_OUTLINE,
                                    on_click=self.eliminar_click,
                                    style=ft.ButtonStyle(
                                        color=ft.Colors.WHITE,
                                        bgcolor=ft.Colors.RED_500,
                                        padding=ft.padding.symmetric(horizontal=24, vertical=12),
                                        shape=ft.RoundedRectangleBorder(radius=8),
                                        elevation=2,
                                    ),
                                    height=48,
                                ),
                                ft.Container(width=16),
                             ], alignment=ft.MainAxisAlignment.CENTER)
                        ]),
                        padding=32,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=16,
                        shadow=ft.BoxShadow(
                            spread_radius=0,
                            blur_radius=8,
                            color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                            offset=ft.Offset(0, 2),
                        ),
                    )
                ]),
                padding=24,
                bgcolor=ft.Colors.GREY_50,
            )
        )
        
        # Pestaña de historial modernizada
        tab_historial = ft.Tab(
            text="Historial",
            icon=ft.Icons.HISTORY_OUTLINED,
            content=ft.Container(
                content=ft.Column([
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.TIMELINE_OUTLINED, size=28, color=ft.Colors.PURPLE_600),
                            ft.Text(
                                "Historial de Verificaciones", 
                                size=24, 
                                weight=ft.FontWeight.W_600,
                                color=ft.Colors.GREY_800,
                            )
                        ], tight=True),
                        margin=ft.margin.only(bottom=24),
                    ),
                    ft.Row([
                        ft.ElevatedButton(
                            "Refrescar Historial",
                            icon=ft.Icons.REFRESH,
                            on_click=lambda _: self.cargar_historial(),
                            style=ft.ButtonStyle(
                                color=ft.Colors.WHITE,
                                bgcolor=ft.Colors.BLUE_600,
                                padding=ft.padding.symmetric(horizontal=24, vertical=12),
                                shape=ft.RoundedRectangleBorder(radius=8),
                                elevation=2,
                            ),
                            height=48,
                        ),
                        ft.Container(width=16),
                        ft.ElevatedButton(
                            "Exportar Historial",
                            icon=ft.Icons.DOWNLOAD_OUTLINED,
                            on_click=self.exportar_historial,
                            style=ft.ButtonStyle(
                                color=ft.Colors.WHITE,
                                bgcolor=ft.Colors.GREEN_600,
                                padding=ft.padding.symmetric(horizontal=24, vertical=12),
                                shape=ft.RoundedRectangleBorder(radius=8),
                                elevation=2,
                            ),
                            height=48,
                        )
                    ], alignment=ft.MainAxisAlignment.CENTER),
                    ft.Container(height=24),
                    self.historial_container
                ]),
                padding=24,
                bgcolor=ft.Colors.GREY_50,
            )
        )
        
        # TabView modernizado
        tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            indicator_color=ft.Colors.BLUE_600,
            label_color=ft.Colors.BLUE_600,
            unselected_label_color=ft.Colors.GREY_500,
            indicator_tab_size=True,
            tabs=[tab_escaneo, tab_crear, tab_buscar, tab_listar, tab_historial],
            expand=True,
            tab_alignment=ft.TabAlignment.CENTER,
        )
        # Agregar el TabView a la página
        page.add(tabs)
        self.cargar_registros()
        self.mostrar_toast("Sistema iniciado correctamente",ft.Colors.GREEN_400)

        
    def inicializar_componentes_gestion(self):
        """Inicializa los componentes para gestión de registros"""
        # Campos para el formulario con diseño moderno
        self.nombre_input = ft.TextField(
            label="Nombre del Estudiante",
            width=400,
            height=56,
            border_radius=12,
            filled=True,
            bgcolor=ft.Colors.GREY_50,
            border_color=ft.Colors.BLUE_200,
            focused_border_color=ft.Colors.BLUE_500,
            label_style=ft.TextStyle(color=ft.Colors.GREY_700, size=14),
            text_style=ft.TextStyle(size=16),
            content_padding=ft.padding.symmetric(horizontal=16, vertical=12),
            prefix_icon=ft.Icons.PERSON_OUTLINE,
        )

        self.cedula_input = ft.TextField(
            label="Número de Cédula",
            width=400,
            height=56,
            border_radius=12,
            filled=True,
            bgcolor=ft.Colors.GREY_50,
            border_color=ft.Colors.BLUE_200,
            focused_border_color=ft.Colors.BLUE_500,
            label_style=ft.TextStyle(color=ft.Colors.GREY_700, size=14),
            text_style=ft.TextStyle(size=16),
            content_padding=ft.padding.symmetric(horizontal=16, vertical=12),
            prefix_icon=ft.Icons.BADGE_OUTLINED,
            keyboard_type=ft.KeyboardType.NUMBER,
        )

        self.especialidad_input = ft.TextField(
            label="Especialidad",
            width=400,
            height=56,
            border_radius=12,
            filled=True,
            bgcolor=ft.Colors.GREY_50,
            border_color=ft.Colors.BLUE_200,
            focused_border_color=ft.Colors.BLUE_500,
            label_style=ft.TextStyle(color=ft.Colors.GREY_700, size=14),
            text_style=ft.TextStyle(size=16),
            content_padding=ft.padding.symmetric(horizontal=16, vertical=12),
            prefix_icon=ft.Icons.SCHOOL_OUTLINED,
        )

        self.anio_input = ft.TextField(
            label="Año",
            width=400,
            height=56,
            border_radius=12,
            filled=True,
            bgcolor=ft.Colors.GREY_50,
            border_color=ft.Colors.BLUE_200,
            focused_border_color=ft.Colors.BLUE_500,
            label_style=ft.TextStyle(color=ft.Colors.GREY_700, size=14),
            text_style=ft.TextStyle(size=16),
            content_padding=ft.padding.symmetric(horizontal=16, vertical=12),
            prefix_icon=ft.Icons.CALENDAR_TODAY_OUTLINED,
            keyboard_type=ft.KeyboardType.NUMBER,
        )

        self.seccion_input = ft.TextField(
            label="Sección",
            width=400,
            height=56,
            border_radius=12,
            filled=True,
            bgcolor=ft.Colors.GREY_50,
            border_color=ft.Colors.BLUE_200,
            focused_border_color=ft.Colors.BLUE_500,
            label_style=ft.TextStyle(color=ft.Colors.GREY_700, size=14),
            text_style=ft.TextStyle(size=16),
            content_padding=ft.padding.symmetric(horizontal=16, vertical=12),
            prefix_icon=ft.Icons.CLASS_OUTLINED,
        )

        self.buscar_cedula_input = ft.TextField(
            label="Buscar por Cédula",
            width=400,
            height=56,
            border_radius=12,
            filled=True,
            bgcolor=ft.Colors.GREY_50,
            border_color=ft.Colors.BLUE_200,
            focused_border_color=ft.Colors.BLUE_500,
            label_style=ft.TextStyle(color=ft.Colors.GREY_700, size=14),
            text_style=ft.TextStyle(size=16),
            content_padding=ft.padding.symmetric(horizontal=16, vertical=12),
            prefix_icon=ft.Icons.SEARCH,
            keyboard_type=ft.KeyboardType.NUMBER,
            suffix_icon=ft.Icons.CLEAR,
        )

        # Contenedor para mostrar los resultados con diseño moderno
        self.resultado_container = ft.Container(
            content=ft.Column([], scroll=ft.ScrollMode.AUTO),
            padding=ft.padding.all(20),
            margin=ft.margin.symmetric(vertical=10),
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=16,
            width=750,
            height=250,
            visible=False,
            bgcolor=ft.Colors.WHITE,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=8,
                color=ft.Colors.with_opacity(0.1, ft.Colors.BLACK),
                offset=ft.Offset(0, 2),
            ),
        )

        # Tabla de historial modernizada
        self.tabla_historial = ft.DataTable(
            columns=[
                ft.DataColumn(
                    ft.Text("Cédula", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Fecha", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Hora", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Becado", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
            ], 
            rows=[],
            border=ft.border.all(1, ft.Colors.GREY_200),
            border_radius=12,
            vertical_lines=ft.border.BorderSide(1, ft.Colors.GREY_100),
            horizontal_lines=ft.border.BorderSide(1, ft.Colors.GREY_100),
            data_row_color={
                ft.ControlState.HOVERED: ft.Colors.BLUE_50,
                ft.ControlState.DEFAULT: ft.Colors.WHITE,
            },
            heading_row_color=ft.Colors.GREY_50,
            heading_row_height=48,
            data_row_min_height=44,
            data_row_max_height=44,
            column_spacing=20,
            show_checkbox_column=False,
            
        )

        # Contenedor para la tabla de historial
        self.historial_container = ft.Container(
            content=self.tabla_historial,
            padding=ft.padding.all(16),
            border_radius=16,
            alignment= ft.alignment.center,
            bgcolor=ft.Colors.WHITE,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=12,
                color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                offset=ft.Offset(0, 4),
            )
        )

        # Tabla para mostrar registros modernizada
        self.tabla_registros = ft.DataTable(
            columns=[
                ft.DataColumn(
                    ft.Text("Nombre", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Cédula", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Especialidad", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Año", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                ),
                ft.DataColumn(
                    ft.Text("Sección", 
                        weight=ft.FontWeight.W_600, 
                        color=ft.Colors.GREY_800,
                        size=14)
                )
            ],
            rows=[],
            border=ft.border.all(1, ft.Colors.GREY_200),
            border_radius=12,
            vertical_lines=ft.border.BorderSide(1, ft.Colors.GREY_100),
            horizontal_lines=ft.border.BorderSide(1, ft.Colors.GREY_100),
            data_row_color={
                ft.ControlState.HOVERED: ft.Colors.BLUE_50,
                ft.ControlState.DEFAULT: ft.Colors.WHITE,
            },
            heading_row_color=ft.Colors.GREY_50,
            heading_row_height=48,
            data_row_min_height=44,
            data_row_max_height=44,
            column_spacing=15,
            show_checkbox_column=False,
        )

        # Contenedor para la tabla de registros
        self.registros_container = ft.Container(
            content=self.tabla_registros,
            padding=ft.padding.all(16),
            alignment= ft.alignment.center,
            border_radius=16,
            bgcolor=ft.Colors.WHITE,
            shadow=ft.BoxShadow(
                spread_radius=1,
                blur_radius=12,
                color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                offset=ft.Offset(0, 4),
            ),
            
        )

        # Mensaje de estado modernizado con diferentes tipos
        self.mensaje_estado = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.INFO_OUTLINE, size=20),
                ft.Text("", size=14, weight=ft.FontWeight.W_500),
            ], alignment=ft.MainAxisAlignment.CENTER),
            padding=ft.padding.symmetric(horizontal=16, vertical=12),
            border_radius=8,
            visible=False,
            animate_opacity=200,
        )
 
    # Métodos para el escáner de cédula
    def on_iniciar_escaneo(self, e):
        """Manejador del evento de clic en el botón Iniciar"""
        if self.escaner is None:
            self.status_text.value = "Estado: Inicializando cámara y OCR..."
            self.status_text.color = ft.Colors.ORANGE
            self.cedula_text.value = "Cédula: Buscando..."
            self.update_ui()
            self.detenido = False
            # Crear y configurar el escáner
            if not self.escaner:
                self.escaner = EscanerCedula(
                    on_cedula_found=self.on_cedula_found,
                    on_scan_failed=self.on_scan_failed,
                    on_frame_update=self.on_frame_update
                )
            if self.escaner:
                print(self.escaner)
            # Iniciar escaneo en un hilo separado
            self.escaner.iniciar_escaneo()
    
    def on_detener_escaneo(self, e):
        """Manejador del evento de clic en el botón Detener"""
        if self.escaner:
            self.detenido = True
            self.escaner.detener_escaneo()
            self.status_text.value = "Estado: Escaneo detenido"
            self.status_text.color = ft.Colors.RED
            self.cedula_text.value = "Cédula: No detectada"
            self.escaner = None
            self.restablecer_imagen()
            
            self.update_ui()
            
    def restablecer_imagen(self):
        """Método auxiliar para restablecer la imagen a su estado original"""
        if self.img_view:
            self.img_view.content.src = "assets/desactivar-camara.png"
            self.img_view.content.src_base64 = None
            self.update_ui()
                
    def cargar_historial(self):
        """Carga todo el historial en la tabla"""
        try:
            historial = gestor.listar_historial_completo()
            
            # Limpiar filas existentes
            self.tabla_historial.rows.clear()
            
            # Agregar nuevas filas
            for entrada in historial:
                self.tabla_historial.rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(entrada['numero_de_cedula'])),
                            ft.DataCell(ft.Text(entrada['dia'])),
                            ft.DataCell(ft.Text(entrada['hora'])),
                            ft.DataCell(ft.Text(entrada['becado'])),
                        ]
                    )
                )
            
            self.page.update()
            self.mostrar_toast(f"Historial cargado: {len(historial)} entradas", ft.Colors.GREEN)
        except Exception as ex:
            self.mostrar_toast(f"Error al cargar historial: {str(ex)}", ft.Colors.RED)
    
    def mostrar_toast(self, mensaje: str, color: str = ft.Colors.GREEN): 
        toast = ft.SnackBar(
            content=ft.Text(mensaje, color=ft.Colors.WHITE),
            bgcolor=color,
            duration=2000,
        )
        self.page.open(toast)
        self.page.update()


    def exportar_historial(self, e):    
        """Exporta el historial a un archivo de Excel con estilos"""
        try:
            

            historial = gestor.listar_historial_completo()

            if not historial:
                self.mostrar_toast("No hay datos en el historial para exportar", ft.Colors.ORANGE)
                return

            fecha_actual = datetime.now().strftime("%m_%d_%Y")
            nombre_archivo = f"historial_verificaciones_{fecha_actual}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial"

            # --- ESTILOS ---
            bold_font = Font(bold=True)
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=16)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))

            # --- TÍTULO ---
            fecha_visible = datetime.now().strftime("%d/%m/%Y")
            titulo = f"Registros - {fecha_visible}"
            ws.merge_cells('A1:D1')
            ws['A1'] = titulo
            ws['A1'].font = title_font
            ws['A1'].alignment = center_align

            # --- ENCABEZADOS ---
            encabezados = ["Cédula", "Fecha", "Hora", "Becado"]
            ws.append([""] * len(encabezados))  # Fila vacía debajo del título
            ws.append(encabezados)

            for col, encabezado in enumerate(encabezados, start=1):
                cell = ws.cell(row=3, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            # --- DATOS ---
            for idx, entrada in enumerate(historial, start=4):
                ws.cell(row=idx, column=1, value=entrada["numero_de_cedula"]).alignment = center_align
                ws.cell(row=idx, column=2, value=entrada["dia"]).alignment = center_align
                ws.cell(row=idx, column=3, value=entrada["hora"]).alignment = center_align
                ws.cell(row=idx, column=4, value=entrada["becado"]).alignment = center_align
                for col in range(1, 5):
                    ws.cell(row=idx, column=col).border = thin_border

            # --- TOTALES Y FECHA ---
            fila_total = len(historial) + 5
            ws[f"A{fila_total}"] = f"Total de entradas: {len(historial)}"
            ws[f"A{fila_total + 1}"] = f"Exportado el: {fecha_visible}"
            ws[f"A{fila_total}"].font = bold_font
            ws[f"A{fila_total + 1}"].font = bold_font

            # --- AJUSTAR ANCHO DE COLUMNAS ---
            for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, max_col=ws.max_column), start=1):
                max_length = 0
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

            # --- CONGELAR ENCABEZADOS ---
            ws.freeze_panes = "A4"

            # --- GUARDAR ---
            import os
            try:
                self.mostrar_toast(f"Historial exportado a: {nombre_archivo}", ft.Colors.GREEN)
            except Exception as e:
                print(f"Error al mostrar el toast: {e}")
            os.makedirs("historiales", exist_ok=True)
            wb.save(f"historiales/{nombre_archivo}")

        except Exception as ex:
            print(ex)
            self.mostrar_toast(f"Error al exportar historial: {str(ex)}", ft.Colors.RED)

    def on_cedula_found(self, cedula):
        """Callback cuando se encuentra una cédula"""
        print(f"[INFO] Mostrando modal para cédula: {cedula}")
        if not self.page:
            print("[ERROR] self.page no está inicializado")
            return
            
        # Actualizar textos en la interfaz
        self.cedula_encontrada = cedula
        self.status_text.value = "Estado: ¡Cédula detectada!"
        self.status_text.color = ft.Colors.GREEN
        self.cedula_text.value = f"Cédula: {cedula}"
        
        # Pausar el escaneo temporalmente
        if self.escaner:
            self.escaner.detener_escaneo()
        
        is_registered = gestor.buscar_por_cedula(cedula)
        already_pass_today = True if len(gestor.buscar_historial_por_cedula(cedula)) > 1 else False
        if is_registered and not already_pass_today:
            # AGREGAR AL HISTORIAL CUANDO SE ESCANEA UNA CÉDULA REGISTRADA
            try:
                gestor.agregar_entrada_historial(cedula, becado="Si")
                print(f"[INFO] Entrada agregada al historial para cédula: {cedula}")
            except Exception as e:
                print(f"[ERROR] Error al agregar al historial: {e}")
            
            # Crear el diálogo modal con todos los datos del estudiante
            modal_dlg = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.VERIFIED_USER, color=ft.Colors.GREEN, size=32),
                    ft.Text("Estudiante Registrado", size=24, weight=ft.FontWeight.BOLD)
                ]),
                content=ft.Container(
                    content=ft.Row([
                        # Columna izquierda - Información del estudiante
                        ft.Column([
                            ft.Card(
                                content=ft.Container(
                                    content=ft.Column(
                                        [
                                            ft.Container(
                                                content=ft.Text(
                                                    "INFORMACIÓN DEL ESTUDIANTE", 
                                                    weight=ft.FontWeight.BOLD, 
                                                    size=18,
                                                    color=ft.Colors.BLUE
                                                ),
                                                alignment=ft.alignment.center,
                                                padding=ft.padding.only(bottom=10)
                                            ),
                                            ft.Divider(),
                                            
                                            # Información organizada en filas
                                            self._create_info_row("Nombre:", is_registered["nombre_estudiante"], ft.Colors.BLACK),
                                            self._create_info_row("Cédula:", is_registered["numero_de_cedula"], ft.Colors.GREEN),
                                            self._create_info_row("Especialidad:", is_registered["especialidad"], ft.Colors.BLACK),
                                            self._create_info_row("Año:", is_registered["año"], ft.Colors.BLACK),
                                            self._create_info_row("Sección:", is_registered["sección"], ft.Colors.BLACK),
                                            self._create_info_row("Código Hash:", is_registered["codigo_hash"][:16] + "...", ft.Colors.BLUE_GREY),
                                            
                                            ft.Divider(),
                                            
                                            # Indicador de registro en historial
                                            ft.Container(
                                                content=ft.Row([
                                                    ft.Icon(ft.Icons.HISTORY, color=ft.Colors.GREEN, size=20),
                                                    ft.Text("Verificación registrada", 
                                                        size=14, 
                                                        color=ft.Colors.GREEN,
                                                        weight=ft.FontWeight.BOLD)
                                                ], 
                                                alignment=ft.MainAxisAlignment.CENTER),
                                                bgcolor=ft.Colors.GREEN_50,
                                                padding=10,
                                                border_radius=8,
                                            ),
                                        ],
                                        spacing=8,
                                        scroll=ft.ScrollMode.AUTO,
                                    ),
                                    padding=20,
                                    border_radius=10,
                                ),
                                elevation=3,
                            ),
                        ], expand=2),
                        
                        # Columna derecha - Imagen del estudiante
                        ft.Column([
                            ft.Card(
                                content=ft.Container(
                                    content=ft.Column([
                                        ft.Text(
                                            "FOTO",
                                            weight=ft.FontWeight.BOLD,
                                            size=16,
                                            color=ft.Colors.BLUE,
                                            text_align=ft.TextAlign.CENTER
                                        ),
                                        ft.Divider(),
                                        ft.Container(
                                            content=ft.Image(
                                                src=f"assets/images/{cedula}.jpg",
                                                width=200,
                                                height=200,
                                                fit=ft.ImageFit.COVER,
                                                border_radius=ft.border_radius.all(10),
                                                error_content=ft.Container(
                                                    content=ft.Column([
                                                        ft.Icon(
                                                            ft.Icons.PERSON,
                                                            size=80,
                                                            color=ft.Colors.GREY_400
                                                        ),
                                                        ft.Text(
                                                            "Sin imagen",
                                                            size=12,
                                                            color=ft.Colors.GREY_600
                                                        )
                                                    ],
                                                    alignment=ft.MainAxisAlignment.CENTER,
                                                    horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                                    width=200,
                                                    height=200,
                                                    bgcolor=ft.Colors.GREY_100,
                                                    border_radius=10,
                                                    alignment=ft.alignment.center
                                                )
                                            ),
                                            alignment=ft.alignment.center,
                                            padding=10
                                        )
                                    ]),
                                    padding=20,
                                    border_radius=10,
                                ),
                                elevation=3,
                            ),
                        ], expand=1),
                    ], spacing=15),
                    width=700,
                    height=400,
                ),
                actions=[
                    ft.Container(
                        content=ft.Row([
                            ft.ElevatedButton(
                                "Continuar Escaneando",
                                icon=ft.Icons.QR_CODE_SCANNER,
                                color=ft.Colors.WHITE,
                                bgcolor=ft.Colors.GREEN,
                                on_click=lambda e: self.page.close(modal_dlg)
                            )
                        ], 
                        alignment=ft.MainAxisAlignment.END,
                        spacing=10),
                        padding=ft.padding.only(top=10)
                    )
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                on_dismiss=lambda e: self.on_detener_escaneo(e),
            )
            sonido = pygame.mixer.Sound('assets/success.mp3')
        elif is_registered and already_pass_today:
            # Mostrar diálogo de ERROR - estudiante ya pasó hoy (NO se agrega al historial)
            modal_dlg = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.BLOCK, color=ft.Colors.RED, size=32),
                    ft.Text("Acceso Denegado", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.RED)
                ]),
                content=ft.Container(
                    content=ft.Card(
                        content=ft.Container(
                            content=ft.Column(
                                [
                                    ft.Icon(
                                        name=ft.Icons.ERROR_OUTLINE,
                                        color=ft.Colors.RED,
                                        size=80,
                                    ),
                                    ft.Text(
                                        "Ya Registrado Hoy",
                                        size=24,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.Colors.RED,
                                        text_align=ft.TextAlign.CENTER,
                                    ),
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Text(
                                                f"Cédula: {cedula}",
                                                size=16,
                                                text_align=ft.TextAlign.CENTER,
                                                color=ft.Colors.BLACK,
                                                weight=ft.FontWeight.BOLD,
                                            ),
                                            ft.Text(
                                                f"Estudiante: {is_registered['nombre_estudiante']}",
                                                size=16,
                                                text_align=ft.TextAlign.CENTER,
                                                color=ft.Colors.BLACK,
                                            ),
                                        ], spacing=5),
                                        padding=ft.padding.only(top=10, bottom=10)
                                    ),
                                    ft.Text(
                                        "Este estudiante ya fue registrado el día de hoy",
                                        size=16,
                                        text_align=ft.TextAlign.CENTER,
                                        color=ft.Colors.GREY_700,
                                    ),
                                    ft.Container(
                                        content=ft.Text(
                                            "No se permite más de un registro por día",
                                            size=14,
                                            text_align=ft.TextAlign.CENTER,
                                            color=ft.Colors.RED,
                                            weight=ft.FontWeight.BOLD,
                                            italic=True
                                        ),
                                        padding=ft.padding.only(top=10),
                                        bgcolor=ft.Colors.RED_50,
                                        border_radius=8,
                                        margin=ft.margin.symmetric(horizontal=20),
                                        alignment=ft.alignment.center,
                                        height=40
                                    )
                                ],
                                spacing=15,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                            ),
                            padding=30,
                            border_radius=10,
                        ),
                        elevation=3,
                    ),
                    width=450,
                    height=350,
                ),
                actions=[
                    ft.Container(
                        content=ft.Row([
                            ft.ElevatedButton(
                                "Intentar de Nuevo",
                                icon=ft.Icons.REFRESH,
                                color=ft.Colors.WHITE,
                                bgcolor=ft.Colors.BLUE,
                                on_click=lambda e: self.page.close(modal_dlg)
                            ),
                            ft.OutlinedButton(
                                "Cerrar",
                                icon=ft.Icons.CLOSE,
                                on_click=lambda e: self.page.close(modal_dlg)
                            ),
                        ], 
                        alignment=ft.MainAxisAlignment.END,
                        spacing=10),
                        padding=ft.padding.only(top=10)
                    )
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                on_dismiss=lambda e: self.page.close(modal_dlg),
            )
            sonido = pygame.mixer.Sound('assets/wrong.mp3')
        else:
            gestor.agregar_entrada_historial(cedula, becado="No")
            # Mostrar diálogo de cédula no registrada (NO se agrega al historial)
            modal_dlg = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED, size=32),
                    ft.Text("Cédula No Registrada", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.RED)
                ]),
                content=ft.Container(
                    content=ft.Card(
                        content=ft.Container(
                            content=ft.Column(
                                [
                                    ft.Icon(
                                        name=ft.Icons.WARNING_ROUNDED,
                                        color=ft.Colors.RED,
                                        size=80,
                                    ),
                                    ft.Text(
                                        "Acceso Denegado",
                                        size=24,
                                        weight=ft.FontWeight.BOLD,
                                        color=ft.Colors.RED,
                                        text_align=ft.TextAlign.CENTER,
                                    ),
                                    ft.Text(
                                        f"La cédula {cedula} no está registrada en el sistema",
                                        size=16,
                                        text_align=ft.TextAlign.CENTER,
                                        color=ft.Colors.GREY_700,
                                    ),
                                    ft.Container(
                                        content=ft.Text(
                                            "Por favor, contacte con el administrador para registrar esta cédula",
                                            size=14,
                                            text_align=ft.TextAlign.CENTER,
                                            color=ft.Colors.GREY_600,
                                            italic=True
                                        ),
                                        padding=ft.padding.only(top=10)
                                    )
                                ],
                                spacing=15,
                                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                            ),
                            padding=30,
                            border_radius=10,
                        ),
                        elevation=3,
                    ),
                    width=400,
                    height=300,
                ),
                actions=[
                    ft.Container(
                        content=ft.Row([
                            ft.ElevatedButton(
                                "Intentar de Nuevo",
                                icon=ft.Icons.REFRESH,
                                color=ft.Colors.WHITE,
                                bgcolor=ft.Colors.BLUE,
                                on_click=lambda e: self.page.close(modal_dlg)
                            ),
                            ft.OutlinedButton(
                                "Cerrar",
                                icon=ft.Icons.CLOSE,
                                on_click=lambda e: self.page.close(modal_dlg)
                            ),
                        ], 
                        alignment=ft.MainAxisAlignment.END,
                        spacing=10),
                        padding=ft.padding.only(top=10)
                    )
                ],
                actions_alignment=ft.MainAxisAlignment.END,
                on_dismiss=lambda e: self.page.close(modal_dlg),
            )
            sonido = pygame.mixer.Sound('assets/wrong.mp3')
        
        self.page.open(modal_dlg)
        sonido.play()

    def _create_info_row(self, label, value, color=ft.Colors.BLACK):
        """Crea una fila de información consistente"""
        return ft.Container(
            content=ft.Row([
                ft.Container(
                    content=ft.Text(label, size=14, weight=ft.FontWeight.W_500, color=ft.Colors.GREY_700),
                    width=120,
                ),
                ft.Container(
                    content=ft.Text(
                        str(value), 
                        size=16, 
                        weight=ft.FontWeight.BOLD,
                        color=color,
                        selectable=True
                    ),
                    expand=True
                )
            ]),
            padding=ft.padding.symmetric(vertical=4),
            border=ft.border.only(bottom=ft.BorderSide(0.5, ft.Colors.GREY_300))
        )
    
    def on_scan_failed(self):
        """Callback cuando no se encuentra una cédula"""
        pass  # No hacemos nada especial aquí para evitar actualizar la UI constantemente
    
    def on_frame_update(self, img_bytes):
        """Callback cuando hay un nuevo frame disponible"""
        if self.img_view and not self.detenido:
            # Convertir bytes a base64 para mostrar en la imagen
            base64_img = base64.b64encode(img_bytes).decode('utf-8')
            self.img_view.content.src_base64 = base64_img
            self.update_ui()
    
    def limpiar_formulario(self):
        """Limpia los campos del formulario"""
        self.nombre_input.value = ""
        self.cedula_input.value = ""
        self.especialidad_input.value = ""
        self.anio_input.value = ""
        self.seccion_input.value = ""
        self.page.update()
    
    def validar_campos_crear(self):
        """Valida que todos los campos obligatorios estén completos"""
        if not self.nombre_input.value or not self.cedula_input.value or not self.especialidad_input.value or not self.anio_input.value or not self.seccion_input.value:
            self.mostrar_toast("Todos los campos son obligatorios", ft.Colors.RED)
            return False
        return True
    
    def crear_click(self, e):
        """Manejador para crear un nuevo registro"""
        if not self.validar_campos_crear():
            return
        
        try:
            resultado = gestor.crear_registro(
                self.nombre_input.value,
                self.cedula_input.value,
                self.especialidad_input.value,
                self.anio_input.value,
                self.seccion_input.value
            )
            
            if resultado:
                self.mostrar_toast("Registro creado exitosamente", ft.Colors.GREEN)
                self.limpiar_formulario()
                self.cargar_registros()
            else:
                self.mostrar_toast("Error al crear el registro", ft.Colors.RED)
        except Exception as ex:
            self.mostrar_toast(f"Error: {str(ex)}", ft.Colors.RED)
    
    def cargar_registros(self):
        """Carga todos los registros en la tabla"""
        try:
            registros = gestor.listar_registros()
            
            # Limpiar filas existentes
            self.tabla_registros.rows.clear()
            
            # Agregar nuevas filas
            for reg in registros:
                self.tabla_registros.rows.append(
                    ft.DataRow(
                        cells=[
                            ft.DataCell(ft.Text(reg['nombre_estudiante'])),
                            ft.DataCell(ft.Text(reg['numero_de_cedula'])),
                            ft.DataCell(ft.Text(reg['especialidad'])),
                            ft.DataCell(ft.Text(reg['año'])),
                            ft.DataCell(ft.Text(reg['sección'])),
                        ]
                    )
                )
            
            self.page.update()
        except Exception as ex:
            self.mostrar_toast(f"Error al cargar registros: {str(ex)}", ft.Colors.RED)
    
    def buscar_click(self, e):
        """Manejador para buscar un registro por cédula"""
        if not self.buscar_cedula_input.value:
            self.mostrar_toast("Ingrese un número de cédula para buscar", ft.Colors.RED)
            return
        
        try:
            resultado = gestor.buscar_por_cedula(self.buscar_cedula_input.value)
            
            self.resultado_container.content.controls.clear()
            
            if resultado:
                self.resultado_container.visible = True
                self.resultado_container.content.controls.append(
                    ft.Column([
                        ft.Text(f"Nombre: {resultado['nombre_estudiante']}", size=16),
                        ft.Text(f"Cédula: {resultado['numero_de_cedula']}", size=16),
                        ft.Text(f"Especialidad: {resultado['especialidad']}", size=16),
                        ft.Text(f"Año: {resultado['año']}", size=16),
                        ft.Text(f"Sección: {resultado['sección']}", size=16),
                        ft.Text(f"Código Hash: {resultado['codigo_hash']}", size=14),
                    ])
                )
                self.mostrar_toast("Registro encontrado", ft.Colors.GREEN)
            else:
                self.resultado_container.visible = True
                self.resultado_container.content.controls.append(
                    ft.Text("No se encontró ningún registro con esa cédula", size=16, color=ft.Colors.RED)
                )
                self.mostrar_toast("No se encontró el registro", ft.Colors.ORANGE)
            
            self.page.update()
        except Exception as ex:
            self.mostrar_toast(f"Error en la búsqueda: {str(ex)}", ft.Colors.RED)
    
    def eliminar_click(self, e):
        """Manejador para eliminar un registro por cédula"""
        if not self.buscar_cedula_input.value:
            self.mostrar_toast("Ingrese un número de cédula para eliminar", ft.Colors.RED)
            return
        
        try:
            resultado = gestor.eliminar_registro(self.buscar_cedula_input.value)
            
            if resultado:
                self.mostrar_toast("Registro eliminado exitosamente", ft.Colors.GREEN)
                self.resultado_container.visible = False
                self.buscar_cedula_input.value = ""
                self.cargar_registros()
            else:
                self.mostrar_toast("No se pudo eliminar el registro", ft.Colors.RED)
            
            self.page.update()
        except Exception as ex:
            self.mostrar_toast(f"Error al eliminar: {str(ex)}", ft.Colors.RED)
    
    def actualizar_click(self, e):
        """Manejador para actualizar un registro existente"""
        if not self.buscar_cedula_input.value:
            self.mostrar_toast("Primero busque un registro para actualizar", ft.Colors.RED)
            return
        
        try:
            datos_actualizados = {}
            
            if self.nombre_input.value:
                datos_actualizados["nombre_estudiante"] = self.nombre_input.value
            if self.especialidad_input.value:
                datos_actualizados["especialidad"] = self.especialidad_input.value
            if self.anio_input.value:
                datos_actualizados["año"] = self.anio_input.value
            if self.seccion_input.value:
                datos_actualizados["sección"] = self.seccion_input.value
            
            if not datos_actualizados:
                self.mostrar_toast("No hay datos para actualizar", ft.Colors.ORANGE)
                return
            
            resultado = gestor.actualizar_registro(self.buscar_cedula_input.value, datos_actualizados)
            
            if resultado:
                self.mostrar_toast("Registro actualizado exitosamente", ft.Colors.GREEN)
                self.limpiar_formulario()
                self.buscar_cedula_input.value = ""
                self.resultado_container.visible = False
                self.cargar_registros()
            else:
                self.mostrar_toast("No se pudo actualizar el registro", ft.Colors.RED)
            
            self.page.update()
        except Exception as ex:
            self.mostrar_toast(f"Error al actualizar: {str(ex)}", ft.Colors.RED)
    
    def update_ui(self):
        """Actualiza componentes específicos de la interfaz de usuario"""
        if self.img_view:
            self.img_view.update()
        if self.status_text:
            self.status_text.update()
        if self.cedula_text:
            self.cedula_text.update()


def main():
    app = CedulaApp()
    # Usar view para asegurar que los diálogos funcionen correctamente
    ft.app(target=app.iniciar_app, view=ft.AppView.FLET_APP)

# Ejecución principal
if __name__ == "__main__":
    main()